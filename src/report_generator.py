import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------
# PATHS
# --------------------------------------------------

BASE_DIR = os.path.dirname(os.path.dirname(__file__))

CSV_FILE = os.path.join(
    BASE_DIR,
    "outputs",
    "screening_results.csv"
)

OUTPUT_FILE = os.path.join(
    BASE_DIR,
    "outputs",
    "screening_report.xlsx"
)


# --------------------------------------------------
# CHECK CSV
# --------------------------------------------------

if not os.path.exists(CSV_FILE):
    print("ERROR: screening_results.csv not found.")
    print(f"Expected location: {CSV_FILE}")
    exit()


# --------------------------------------------------
# CREATE WORKBOOK
# --------------------------------------------------

workbook = Workbook()

sheet = workbook.active
sheet.title = "Screening Results"


# --------------------------------------------------
# READ CSV
# --------------------------------------------------

with open(
    CSV_FILE,
    "r",
    encoding="utf-8-sig",
    newline=""
) as file:

    reader = csv.reader(file)

    rows = list(reader)


# --------------------------------------------------
# WRITE DATA
# --------------------------------------------------

for row in rows:
    sheet.append(row)


# --------------------------------------------------
# HEADER STYLE
# --------------------------------------------------

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


for cell in sheet[1]:

    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    cell.border = thin_border


# --------------------------------------------------
# FORMAT DATA
# --------------------------------------------------

for row in sheet.iter_rows(
    min_row=2,
    max_row=sheet.max_row
):

    for cell in row:

        cell.border = thin_border

        cell.alignment = Alignment(
            vertical="center"
        )


# --------------------------------------------------
# CENTER SELECTED COLUMNS
# --------------------------------------------------

for column in ["A", "C", "D", "E", "F", "G", "H", "J"]:

    for cell in sheet[column][1:]:

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


# --------------------------------------------------
# COLUMN WIDTHS
# --------------------------------------------------

widths = {
    "A": 8,
    "B": 22,
    "C": 18,
    "D": 15,
    "E": 15,
    "F": 18,
    "G": 15,
    "H": 15,
    "I": 75,
    "J": 18
}

for column, width in widths.items():

    sheet.column_dimensions[
        column
    ].width = width


# --------------------------------------------------
# FREEZE HEADER
# --------------------------------------------------

sheet.freeze_panes = "A2"


# --------------------------------------------------
# AUTO FILTER
# --------------------------------------------------

sheet.auto_filter.ref = sheet.dimensions


# --------------------------------------------------
# ADD SUMMARY SHEET
# --------------------------------------------------

summary = workbook.create_sheet(
    "Summary"
)

summary["A1"] = "RESUME SCREENING SUMMARY"

summary["A1"].font = Font(
    bold=True,
    size=16
)

summary["A3"] = "Total Candidates"
summary["B3"] = len(rows) - 1

summary["A4"] = "Strong Match"
summary["B4"] = '=COUNTIF(\'Screening Results\'!J:J,"Strong Match")'

summary["A5"] = "Consider"
summary["B5"] = '=COUNTIF(\'Screening Results\'!J:J,"Consider")'

summary["A6"] = "Maybe"
summary["B6"] = '=COUNTIF(\'Screening Results\'!J:J,"Maybe")'

summary["A7"] = "Reject"
summary["B7"] = '=COUNTIF(\'Screening Results\'!J:J,"Reject")'


summary["A9"] = "Top Candidate"

summary["B9"] = (
    "='Screening Results'!B2"
)

summary["A10"] = "Top Candidate Score"

summary["B10"] = (
    "='Screening Results'!D2"
)


# --------------------------------------------------
# SUMMARY FORMATTING
# --------------------------------------------------

for row in summary.iter_rows(
    min_row=3,
    max_row=10,
    min_col=1,
    max_col=2
):

    for cell in row:

        cell.border = thin_border
        cell.alignment = Alignment(
            vertical="center"
        )


for cell in summary["A"]:

    if cell.row >= 3:

        cell.font = Font(
            bold=True
        )


summary.column_dimensions["A"].width = 25
summary.column_dimensions["B"].width = 30


# --------------------------------------------------
# SAVE WORKBOOK
# --------------------------------------------------

workbook.save(
    OUTPUT_FILE
)


# --------------------------------------------------
# SUCCESS MESSAGE
# --------------------------------------------------

print()
print("=" * 70)
print("EXCEL REPORT CREATED")
print("=" * 70)

print(
    f"Excel file: {OUTPUT_FILE}"
)

print()
print("Report contains:")
print("1. Screening Results")
print("2. Summary")
print("3. Candidate rankings")
print("4. Scores")
print("5. Matched skills")
print("6. Decisions")

print()
print("Excel report generated successfully.")